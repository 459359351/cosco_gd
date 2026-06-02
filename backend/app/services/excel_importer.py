"""Excel 周报导入解析器 — 仅从主 Sheet + 同比 Sheet 提取全部数据。

主 Sheet 结构（以"2026年第20周"为例）：
  R1-R3        表头
  R4-R9        自营机构行（[24-28]列含累计数据）
  R10          自营总计
  R11-R34      外包机构行
  R35          外包总计
  R36          自营外包总计
  R37-R41      占比/汇总
  R42          "网点明细报表" 标题
  R43-R44      网点明细表头
  R45-R182     各网点明细（[1]经营单位 [2]网点名 + 箱量/产值等）
"""

from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.repair import (
    ExcelUploadLog,
    RepairNetworkSite,
    RepairWeeklyCumulative,
    RepairWeeklyOrg,
    RepairWeeklySiteDetail,
    RepairWeeklySummary,
)

# ---------------------------------------------------------------------------
# 常量
# ---------------------------------------------------------------------------

SELF_ORG_NAMES = {
    "深圳运营中心",
    "广州运营中心",
    "珠三角运营中心",
    "粤西运营中心",
    "北部湾运营中心",
    "南岗分公司",
}

TOTAL_PATTERNS = {
    "combined_total": ("自营外包总计",),
    "self_total": ("自营总计",),
    "outsourced_total": ("外包总计",),
}


def _safe_float(val) -> float:
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _safe_int(val) -> int:
    if val is None:
        return 0
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return 0


def _cell(ws, row: int, col: int):
    v = ws.cell(row=row, column=col).value
    return None if v is None or (isinstance(v, str) and v.strip() == "") else v


def _classify_company(name: str) -> str:
    return "self" if name in SELF_ORG_NAMES else "outsourced"


def _make_code(name: str) -> str:
    return re.sub(r"[^\w一-鿿]", "", name)


_SHEET_WEEK_RE = re.compile(r"(\d{4})\s*年?\s*第?\s*(\d{1,2})\s*周")


def parse_sheet_name(name: str) -> dict:
    m = _SHEET_WEEK_RE.search(name)
    info: dict = {"detected_year": None, "detected_week": None, "suggested_role": None}
    if m:
        info["detected_year"] = int(m.group(1))
        info["detected_week"] = int(m.group(2))
    return info


# ---------------------------------------------------------------------------
# 列映射：中远海 cosco / 第三方 thirdparty
# ---------------------------------------------------------------------------

# (qty, qty_wow, rev, rev_wow, unit_price, move_fee, total_qty, total_rev, per_capita, staff)
COLS_COSCO = (3, 4, 5, 6, 7, 9, 20, 21, 22, 23)
COLS_THIRD = (10, 11, 12, 13, 14, 16, 20, 21, 22, 23)


# ---------------------------------------------------------------------------
# 主解析器
# ---------------------------------------------------------------------------


class WeeklyExcelImporter:
    def __init__(self, file_path: str | Path, year: int, week: int):
        self.file_path = Path(file_path)
        self.year = year
        self.week = week
        self.wb = load_workbook(str(self.file_path), data_only=True)
        self.row_counts: dict[str, int] = {}
        self.site_registry: dict[str, dict] = {}

    # -- 公共入口 --

    def parse_all(self) -> dict:
        ws = self._find_main_sheet()
        if ws is None:
            return {"orgs": [], "summaries": [], "sites": [], "site_details": [], "cumulative": []}

        # 先解析网点明细区域以建立 site_registry
        sites = self._parse_site_zone(ws)
        # 再解析机构汇总区域
        orgs = self._parse_org_zone(ws)
        summaries = self._parse_summary_zone(ws)
        cumulative = self._parse_cumulative_sheet()

        return {
            "orgs": orgs,
            "summaries": summaries,
            "sites": sites,
            "site_details": self._build_site_details(ws),
            "cumulative": cumulative,
        }

    def parse_yoy(self, yoy_sheet_name: str) -> dict:
        prev_year, prev_week = self._get_yoy_year_week()
        if yoy_sheet_name not in self.wb.sheetnames:
            return {"yoy_orgs": [], "yoy_summaries": []}
        ws = self.wb[yoy_sheet_name]
        return {
            "yoy_orgs": self._parse_org_zone(ws, prev_year, prev_week),
            "yoy_summaries": self._parse_summary_zone(ws, prev_year, prev_week),
        }

    # Sheet 名字中属于辅助/汇总表的过滤词（不应作为主数据或同比 sheet）
    _SKIP_SHEET_KEYWORDS = ("纯", "洗", "新")

    def preview_sheets(self) -> list[dict]:
        sheets: list[dict] = []
        prev_year, prev_week = self._get_yoy_year_week()
        for name in self.wb.sheetnames:
            info = parse_sheet_name(name)
            is_skip = any(kw in name for kw in self._SKIP_SHEET_KEYWORDS)
            if not is_skip and info["detected_year"] == self.year and info["detected_week"] == self.week:
                info["suggested_role"] = "current"
            elif not is_skip and info["detected_year"] == prev_year and info["detected_week"] == prev_week:
                info["suggested_role"] = "yoy"
            sheets.append({"name": name, **info})
        return sheets

    # -- Sheet 定位 --

    def _find_main_sheet(self):
        target_year = str(self.year)
        week_str = str(self.week)
        for name in self.wb.sheetnames:
            if target_year in name and (f"第{week_str}周" in name or f"{week_str}周" in name):
                if "新" not in name and "洗" not in name and "纯" not in name:
                    return self.wb[name]
        return None

    # -- 区域1：机构行 (R4 ~ "网点明细" 之前) --

    def _parse_org_zone(self, ws, target_year: int = None, target_week: int = None) -> list[dict]:
        ty = target_year or self.year
        tw = target_week or self.week
        records: list[dict] = []
        current_type = ""

        for row_idx in range(4, ws.max_row + 1):
            a_val = _cell(ws, row_idx, 1)
            if a_val and "网点明细" in str(a_val):
                break

            b_val = _cell(ws, row_idx, 2)
            if not b_val:
                continue

            a_str = str(a_val).strip() if a_val else ""
            if "自营" in a_str and "合作" not in a_str and "外包" not in a_str:
                current_type = "self"
            elif "合作供应商" in a_str:
                current_type = "outsourced"

            b_str = str(b_val).strip()
            if not b_str or b_str in ("部门/单位", "环比增减"):
                continue
            if any(kw in b_str for kw in ("自营总计", "外包总计", "自营外包总计")):
                continue

            org_name = b_str
            org_code = _make_code(org_name)
            company_type = current_type if current_type else _classify_company(org_name)

            for cust_type, cols in [("cosco", COLS_COSCO), ("thirdparty", COLS_THIRD)]:
                qty_c, wow_c, rev_c, rwow_c, unit_c, move_c, tqty_c, trev_c, pcap_c, staff_c = cols
                records.append({
                    "year": ty, "week": tw,
                    "company_type": company_type,
                    "org_name": org_name, "org_code": org_code,
                    "customer_type": cust_type,
                    "container_qty": _safe_int(_cell(ws, row_idx, qty_c)),
                    "qty_wow_change": _safe_int(_cell(ws, row_idx, wow_c)),
                    "revenue": _safe_float(_cell(ws, row_idx, rev_c)),
                    "rev_wow_change": _safe_float(_cell(ws, row_idx, rwow_c)),
                    "unit_price": _safe_float(_cell(ws, row_idx, unit_c)),
                    "move_fee": _safe_float(_cell(ws, row_idx, move_c)),
                    "total_qty": _safe_int(_cell(ws, row_idx, tqty_c)),
                    "total_revenue": _safe_float(_cell(ws, row_idx, trev_c)),
                    "per_capita": _safe_float(_cell(ws, row_idx, pcap_c)),
                    "staff_count": _safe_int(_cell(ws, row_idx, staff_c)),
                })

        self.row_counts["orgs"] = len(records)
        return records

    # -- 区域2：汇总行（自营总计/外包总计/自营外包总计）--

    def _parse_summary_zone(self, ws, target_year: int = None, target_week: int = None) -> list[dict]:
        ty = target_year or self.year
        tw = target_week or self.week
        records: list[dict] = []

        for row_idx in range(4, ws.max_row + 1):
            a_val = _cell(ws, row_idx, 1)
            if a_val and "网点明细" in str(a_val):
                break

            b_val = _cell(ws, row_idx, 2)
            label = ""
            if b_val:
                label = str(b_val).strip()
            elif a_val:
                label = str(a_val).strip()
            if not label:
                continue

            matched_type = None
            for stype, patterns in TOTAL_PATTERNS.items():
                if any(p in label for p in patterns):
                    matched_type = stype
                    break
            if not matched_type:
                continue

            for cust_type, cols in [("cosco", COLS_COSCO), ("thirdparty", COLS_THIRD)]:
                qty_c, wow_c, rev_c, rwow_c, unit_c, move_c, tqty_c, trev_c, pcap_c, staff_c = cols
                records.append({
                    "year": ty, "week": tw,
                    "summary_type": matched_type,
                    "customer_type": cust_type,
                    "container_qty": _safe_int(_cell(ws, row_idx, qty_c)),
                    "qty_wow_change": _safe_int(_cell(ws, row_idx, wow_c)),
                    "revenue": _safe_float(_cell(ws, row_idx, rev_c)),
                    "rev_wow_change": _safe_float(_cell(ws, row_idx, rwow_c)),
                    "unit_price": _safe_float(_cell(ws, row_idx, unit_c)),
                    "move_fee": _safe_float(_cell(ws, row_idx, move_c)),
                    "total_qty": _safe_int(_cell(ws, row_idx, tqty_c)),
                    "total_revenue": _safe_float(_cell(ws, row_idx, trev_c)),
                    "per_capita": _safe_float(_cell(ws, row_idx, pcap_c)),
                    "staff_count": _safe_int(_cell(ws, row_idx, staff_c)),
                })

        self.row_counts["summaries"] = len(records)
        return records

    # -- 区域3：网点明细 (R42"网点明细报表" 之后) --

    def _parse_site_zone(self, ws) -> list[dict]:
        # 定位"网点明细报表"
        start_row = None
        for row_idx in range(1, ws.max_row + 1):
            v = _cell(ws, row_idx, 1)
            if v and "网点明细" in str(v):
                start_row = row_idx
                break
        if start_row is None:
            return []

        records: list[dict] = []
        for row_idx in range(start_row + 2, ws.max_row + 1):
            parent_name = _cell(ws, row_idx, 1)
            site_name = _cell(ws, row_idx, 2)
            if not parent_name or not site_name:
                continue

            parent_name = str(parent_name).strip()
            site_name = str(site_name).strip()

            # 如果 col1 是表头或非数据行则跳过
            if parent_name in ("部门/单位", "业务类型") or "合计" in parent_name:
                continue

            company_type = _classify_company(parent_name)
            code = _make_code(f"{parent_name}_{site_name}")
            parent_code = _make_code(parent_name)

            self.site_registry[site_name] = {
                "company_type": company_type,
                "parent_name": parent_name,
                "parent_code": parent_code,
            }

            records.append({
                "name": site_name,
                "code": code,
                "company_type": company_type,
                "parent_name": parent_name,
                "parent_code": parent_code,
            })

        self.row_counts["sites"] = len(records)
        return records

    # -- 区域3b：网点明细 → RepairWeeklySiteDetail --

    def _build_site_details(self, ws) -> list[dict]:
        # 定位"网点明细报表"
        start_row = None
        for row_idx in range(1, ws.max_row + 1):
            v = _cell(ws, row_idx, 1)
            if v and "网点明细" in str(v):
                start_row = row_idx
                break
        if start_row is None:
            return []

        records: list[dict] = []
        for row_idx in range(start_row + 2, ws.max_row + 1):
            parent_name = _cell(ws, row_idx, 1)
            site_name = _cell(ws, row_idx, 2)
            if not parent_name or not site_name:
                continue

            parent_name = str(parent_name).strip()
            site_name = str(site_name).strip()
            if parent_name in ("部门/单位", "业务类型") or "合计" in parent_name:
                continue

            parent_info = self._lookup_parent(site_name)

            # 按中远海/第三方分别生成记录
            cosco_qty = _safe_int(_cell(ws, row_idx, 3))
            cosco_rev = _safe_float(_cell(ws, row_idx, 5))
            third_qty = _safe_int(_cell(ws, row_idx, 10))
            third_rev = _safe_float(_cell(ws, row_idx, 12))

            if cosco_qty > 0 or cosco_rev > 0:
                records.append({
                    "year": self.year, "week": self.week,
                    "site_name": site_name,
                    "company_name": parent_name,
                    "company_type": parent_info["company_type"],
                    "container_class": "干箱",
                    "customer_name": "中远海",
                    "repair_qty": cosco_qty,
                    "approved_amount": cosco_rev,
                    "move_fee": _safe_float(_cell(ws, row_idx, 9)),
                })

            if third_qty > 0 or third_rev > 0:
                records.append({
                    "year": self.year, "week": self.week,
                    "site_name": site_name,
                    "company_name": parent_name,
                    "company_type": parent_info["company_type"],
                    "container_class": "干箱",
                    "customer_name": "第三方",
                    "repair_qty": third_qty,
                    "approved_amount": third_rev,
                    "move_fee": 0.0,
                })

        self.row_counts["site_details"] = len(records)
        return records

    # -- 区域4：累计数据（从"自营周完成情况" Sheet 提取）--

    def _find_cumulative_sheet(self):
        """定位 '自营NNNN周完成情况' Sheet。"""
        for name in self.wb.sheetnames:
            if "周完成情况" in name or "完成情况" in name:
                return self.wb[name]
        return None

    def _parse_cumulative_sheet(self) -> list[dict]:
        ws = self._find_cumulative_sheet()
        if ws is None:
            # 回退到从主 Sheet 自营行提取（精度较低）
            return self._parse_cumulative_from_main()

        records: list[dict] = []
        data_start = None
        # 找表头行（含 "周累计箱量" 或 "周累计"）
        for row_idx in range(1, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                v = _cell(ws, row_idx, col_idx)
                if v and "累计箱量" in str(v):
                    data_start = row_idx + 1
                    break
            if data_start:
                break

        if data_start is None:
            return self._parse_cumulative_from_main()

        # 列映射：[1]机构名 [2]周累计箱量 [3]周累计量同比 [4]周累计产值(万) [5]周累计产值同比(万)
        # 只读第一个连续数据块（自营机构 + 自营总计），遇到空行即停
        for row_idx in range(data_start, ws.max_row + 1):
            name_val = _cell(ws, row_idx, 1)
            if not name_val:
                if records:
                    break  # 已有数据后遇到空行 → 第一段结束
                continue
            name_str = str(name_val).strip()
            if not name_str or name_str in ("部门/单位", "环比增减"):
                continue

            cum_qty = _safe_int(_cell(ws, row_idx, 2))
            cum_qty_yoy = _safe_int(_cell(ws, row_idx, 3))
            cum_rev = _safe_float(_cell(ws, row_idx, 4))
            cum_rev_yoy = _safe_float(_cell(ws, row_idx, 5))

            if cum_qty == 0 and cum_rev == 0:
                continue

            records.append({
                "year": self.year, "week": self.week,
                "org_name": name_str, "org_code": _make_code(name_str),
                "cum_qty": cum_qty,
                "cum_qty_yoy": cum_qty_yoy,
                "cum_revenue": cum_rev,
                "cum_revenue_yoy": cum_rev_yoy,
            })

        self.row_counts["cumulative"] = len(records)
        return records

    def _parse_cumulative_from_main(self) -> list[dict]:
        """回退方案：从主 Sheet 自营行 [27] 列取累计箱量。"""
        ws = self._find_main_sheet()
        if ws is None:
            return []

        records: list[dict] = []
        in_self_zone = False

        for row_idx in range(4, ws.max_row + 1):
            a_val = _cell(ws, row_idx, 1)
            if a_val and "网点明细" in str(a_val):
                break

            a_str = str(a_val).strip() if a_val else ""
            if "自营" in a_str and "合作" not in a_str and "外包" not in a_str:
                in_self_zone = True
            elif "合作供应商" in a_str:
                break

            if not in_self_zone:
                continue

            b_val = _cell(ws, row_idx, 2)
            if not b_val:
                continue
            b_str = str(b_val).strip()
            if not b_str or b_str in ("部门/单位", "环比增减"):
                continue
            if any(kw in b_str for kw in ("自营总计", "外包总计", "自营外包总计")):
                continue

            org_name = b_str
            cum_qty = _safe_int(_cell(ws, row_idx, 27))
            if cum_qty > 0:
                records.append({
                    "year": self.year, "week": self.week,
                    "org_name": org_name, "org_code": _make_code(org_name),
                    "cum_qty": cum_qty, "cum_qty_yoy": 0,
                    "cum_revenue": 0.0, "cum_revenue_yoy": 0.0,
                })

        self.row_counts["cumulative"] = len(records)
        return records

    def _lookup_parent(self, site_name: str) -> dict:
        if site_name in self.site_registry:
            return self.site_registry[site_name]
        for key, val in self.site_registry.items():
            if site_name in key or key in site_name:
                return val
        return {"company_type": "outsourced", "parent_name": "", "parent_code": ""}

    def _get_yoy_year_week(self) -> tuple[int, int]:
        return self.year - 1, self.week - 1


# ---------------------------------------------------------------------------
# 数据库写入
# ---------------------------------------------------------------------------


async def import_weekly_excel(
    db: AsyncSession,
    file_path: str | Path,
    year: int,
    week: int,
    filename: str = "",
    yoy_sheet: str | None = None,
) -> dict:
    log = ExcelUploadLog(
        year=year,
        week=week,
        filename=filename,
        status="pending",
        uploaded_at=datetime.now(),
    )
    db.add(log)
    await db.flush()

    try:
        importer = WeeklyExcelImporter(file_path, year, week)
        data = importer.parse_all()

        # 清除旧数据（幂等）
        for table in [
            RepairWeeklyOrg,
            RepairWeeklySummary,
            RepairWeeklySiteDetail,
            RepairWeeklyCumulative,
        ]:
            await db.execute(
                delete(table).where(table.year == year, table.week == week)
            )

        # 网点主数据 upsert
        for s in data["sites"]:
            existing = await db.execute(
                select(RepairNetworkSite).where(
                    RepairNetworkSite.code == s["code"]
                )
            )
            obj = existing.scalar_one_or_none()
            if obj:
                obj.name = s["name"]
                obj.company_type = s["company_type"]
                obj.parent_name = s["parent_name"]
                obj.parent_code = s["parent_code"]
            else:
                db.add(RepairNetworkSite(**s))

        for r in data["orgs"]:
            db.add(RepairWeeklyOrg(**r))
        for r in data["summaries"]:
            db.add(RepairWeeklySummary(**r))
        for r in data["site_details"]:
            db.add(RepairWeeklySiteDetail(**r))
        for r in data["cumulative"]:
            db.add(RepairWeeklyCumulative(**r))

        # 同比数据导入
        yoy_counts: dict[str, int] = {}
        if yoy_sheet:
            prev_year, prev_week = year - 1, week - 1
            yoy_data = importer.parse_yoy(yoy_sheet)

            for table in [RepairWeeklyOrg, RepairWeeklySummary]:
                await db.execute(
                    delete(table).where(table.year == prev_year, table.week == prev_week)
                )

            for r in yoy_data["yoy_orgs"]:
                db.add(RepairWeeklyOrg(**r))
            for r in yoy_data["yoy_summaries"]:
                db.add(RepairWeeklySummary(**r))

            yoy_counts = {
                "yoy_orgs": len(yoy_data["yoy_orgs"]),
                "yoy_summaries": len(yoy_data["yoy_summaries"]),
                "yoy_year": prev_year,
                "yoy_week": prev_week,
            }

        log.status = "success"
        log.row_counts = json.dumps({**importer.row_counts, **yoy_counts}, ensure_ascii=False)
        log.processed_at = datetime.now()
        await db.flush()

        return {
            "status": "success",
            "year": year,
            "week": week,
            "row_counts": {**importer.row_counts, **yoy_counts},
        }

    except Exception as exc:
        log.status = "error"
        log.error_message = str(exc)
        log.processed_at = datetime.now()
        await db.flush()
        raise
