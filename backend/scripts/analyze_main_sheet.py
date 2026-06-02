"""分析主 Sheet 的完整结构，找出所有可提取的数据区域。

用法: python scripts/analyze_main_sheet.py <excel文件路径>
"""

import re
import sys
from pathlib import Path

from openpyxl import load_workbook


def _cell(ws, row: int, col: int):
    v = ws.cell(row=row, column=col).value
    return None if v is None or (isinstance(v, str) and v.strip() == "") else v


def analyze_main_sheet(path: str, year: int = 2026, week: int = 20):
    wb = load_workbook(path, data_only=True)
    print(f"工作簿包含 {len(wb.sheetnames)} 个 Sheet:")
    for name in wb.sheetnames:
        print(f"  - {name}")
    print()

    # 定位主 Sheet
    target_year = str(year)
    week_str = str(week)
    main_ws = None
    main_name = ""
    for name in wb.sheetnames:
        if target_year in name and (f"第{week_str}周" in name or f"{week_str}周" in name):
            if "新" not in name and "洗" not in name and "纯" not in name:
                main_ws = wb[name]
                main_name = name
                break

    if not main_ws:
        print(f"未找到 {year}年第{week}周 的主 Sheet")
        return

    print(f"主 Sheet: \"{main_name}\"")
    print(f"行数: {main_ws.max_row}, 列数: {main_ws.max_column}")
    print()

    # 逐行扫描，标记各区域
    print("=" * 80)
    print("主 Sheet 完整内容（非空行）")
    print("=" * 80)

    zones = []
    current_zone = "header"
    zone_start = 1

    for row_idx in range(1, main_ws.max_row + 1):
        a_val = _cell(main_ws, row_idx, 1)
        b_val = _cell(main_ws, row_idx, 2)

        # 检测区域转换
        if a_val:
            a_str = str(a_val).strip()
            if "自营" in a_str and "合作" not in a_str and "外包" not in a_str:
                if current_zone != "self":
                    zones.append((current_zone, zone_start, row_idx - 1))
                    current_zone = "self"
                    zone_start = row_idx
            elif "合作供应商" in a_str:
                if current_zone != "outsourced":
                    zones.append((current_zone, zone_start, row_idx - 1))
                    current_zone = "outsourced"
                    zone_start = row_idx
            elif "全部干箱" in a_str:
                zones.append((current_zone, zone_start, row_idx - 1))
                current_zone = "summary"
                zone_start = row_idx
            elif "网点明细" in a_str:
                zones.append((current_zone, zone_start, row_idx - 1))
                current_zone = "site_detail"
                zone_start = row_idx
            elif "累计" in a_str or "周完成" in a_str or "完成情况" in a_str:
                zones.append((current_zone, zone_start, row_idx - 1))
                current_zone = "cumulative"
                zone_start = row_idx

        # 收集非空单元格
        cells = []
        for col_idx in range(1, min(main_ws.max_column + 1, 30)):
            v = _cell(main_ws, row_idx, col_idx)
            if v is not None:
                cells.append((col_idx, v))

        if cells:
            row_desc = "  ".join(f"[{c[0]}]{c[1]}" for c in cells)
            print(f"  R{row_idx:3d} | {row_desc}")

    zones.append((current_zone, zone_start, main_ws.max_row))

    # 汇总区域
    print()
    print("=" * 80)
    print("区域汇总")
    print("=" * 80)
    for zone_name, start, end in zones:
        print(f"  {zone_name:15s} | 行 {start:3d} ~ {end:3d}")

    # 重点分析"网点明细"区域
    print()
    print("=" * 80)
    print("网点明细区域详情（如果存在）")
    print("=" * 80)
    site_start = None
    for row_idx in range(1, main_ws.max_row + 1):
        a_val = _cell(main_ws, row_idx, 1)
        if a_val and "网点明细" in str(a_val):
            site_start = row_idx
            break

    if site_start:
        # 打印表头
        for r in range(site_start, min(site_start + 40, main_ws.max_row + 1)):
            cells = []
            for col_idx in range(1, min(main_ws.max_column + 1, 30)):
                v = _cell(main_ws, r, col_idx)
                if v is not None:
                    cells.append((col_idx, v))
            if cells:
                row_desc = "  ".join(f"[{c[0]}]{c[1]}" for c in cells)
                print(f"  R{r:3d} | {row_desc}")
            else:
                print(f"  R{r:3d} | (空行)")
    else:
        print("  主 Sheet 中未找到'网点明细'标记")

    # 重点分析"累计"相关区域
    print()
    print("=" * 80)
    print("累计/周完成区域详情（如果存在）")
    print("=" * 80)
    cum_keywords = ("累计", "周完成", "完成情况", "周累计")
    cum_found = False
    for row_idx in range(1, main_ws.max_row + 1):
        for col_idx in range(1, min(main_ws.max_column + 1, 10)):
            v = _cell(main_ws, row_idx, col_idx)
            if v and any(kw in str(v) for kw in cum_keywords):
                cum_found = True
                # 打印该行及后续数据
                for r in range(row_idx, min(row_idx + 15, main_ws.max_row + 1)):
                    cells = []
                    for c in range(1, min(main_ws.max_column + 1, 30)):
                        val = _cell(main_ws, r, c)
                        if val is not None:
                            cells.append((c, val))
                    if cells:
                        row_desc = "  ".join(f"[{c[0]}]{c[1]}" for c in cells)
                        print(f"  R{r:3d} | {row_desc}")
                    else:
                        break
                print()
                break
        if cum_found:
            break
    if not cum_found:
        print("  主 Sheet 中未找到累计/周完成相关内容")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python scripts/analyze_main_sheet.py <excel文件路径> [year] [week]")
        sys.exit(1)
    excel_path = sys.argv[1]
    y = int(sys.argv[2]) if len(sys.argv) > 2 else 2026
    w = int(sys.argv[3]) if len(sys.argv) > 3 else 20
    analyze_main_sheet(excel_path, y, w)
