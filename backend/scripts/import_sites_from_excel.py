"""用 Excel（广州网点地址（已整理）.xlsx）作为最终地理数据源，
重新地理编码并更新 repair_network_sites 和 yards 两表。

用法: 在 backend/ 目录下运行:
    python scripts/import_sites_from_excel.py
"""

import asyncio
import json
import os
import re
import sys
from pathlib import Path

import aiohttp
import asyncpg
from openpyxl import load_workbook

# ---- 配置 ----
# Excel 路径：脚本在 cosco_gd/backend/scripts/，Excel 在 财商驾驶舱/（cosco_gd 的上级）
_SCRIPT_DIR = Path(__file__).resolve().parent
EXCEL_PATH = _SCRIPT_DIR.parent.parent.parent / "广州网点地址（已整理）.xlsx"
if not EXCEL_PATH.exists():
    EXCEL_PATH = Path(os.getcwd()).parent.parent / "广州网点地址（已整理）.xlsx"
DB_URL = "postgresql://postgres:postgres@localhost:5432/cosco_gd"
AMAP_KEY = "26d9ac20e8f589851bbd5b17c3d4cff4"
OUTPUT_DIR = Path(__file__).resolve().parent.parent  # backend/

BATCH_DELAY = 0.15  # 高德 API 调用间隔（秒）
MAX_RETRIES = 3
RETRY_DELAY = [1, 3, 5]

# 中国坐标合理范围
LNG_MIN, LNG_MAX = 73, 135
LAT_MIN, LAT_MAX = 18, 54


def normalize_province(raw: str) -> str:
    """从高德返回省份中提取简写."""
    for kw in ["广东", "广西", "云南", "贵州", "海南", "福建"]:
        if kw in raw:
            return kw
    return raw.strip().rstrip("省自治区市")


def normalize_city(raw: str) -> str:
    """去除省市后缀."""
    for kw in ["壮族自治区", "自治区", "省", "市"]:
        raw = raw.replace(kw, "")
    return raw.strip()


def make_code(text: str) -> str:
    return re.sub(r"[^\w一-鿿]", "", text)


def extract_city_hint(address: str) -> str:
    """从地址字符串中提取城市作为地理编码提示."""
    for kw in [
        "广州市", "深圳市", "东莞市", "佛山市", "珠海市", "中山市", "江门市",
        "惠州市", "肇庆市", "清远市", "汕头市", "湛江市", "茂名市", "云浮市",
        "揭阳市", "汕尾市", "阳江市",
        "南宁市", "钦州市", "防城港市", "北海市", "梧州市", "贺州市", "百色市",
        "贵港市", "来宾市", "玉林市",
        "昆明市",
        "广州", "深圳", "东莞", "佛山", "珠海", "中山", "江门",
        "惠州", "肇庆", "清远", "汕头", "湛江", "茂名", "云浮",
        "揭阳", "汕尾", "阳江",
        "南宁", "钦州", "防城港", "北海", "梧州", "贺州", "百色",
        "贵港", "来宾", "玉林",
        "昆明",
    ]:
        if kw in address:
            return kw.rstrip("市")
    return ""


def parse_excel(filepath: str) -> tuple[list[dict], list[dict]]:
    """解析 Excel，返回 (sheet1_records, sheet2_records)."""
    wb = load_workbook(filepath, data_only=True)
    ws1 = wb["Sheet1"]
    ws2 = wb["Sheet2"]

    # ---- Sheet1 ----
    sheet1 = []
    current_dept = ""
    for row in ws1.iter_rows(min_row=2, values_only=True):
        dept = str(row[0]).strip() if row[0] else ""
        name = str(row[1]).strip() if row[1] else ""
        code = str(row[2]).strip() if row[2] else ""
        site_type = str(row[3]).strip() if row[3] else ""
        biz_partner = str(row[4]).strip() if row[4] else ""
        contact = str(row[5]).strip() if row[5] else ""
        phone = str(row[6]).strip() if row[6] else ""
        address = str(row[7]).strip() if row[7] else ""

        if not name:
            continue

        if dept:
            current_dept = dept

        sheet1.append({
            "parent_name": current_dept,
            "name": name,
            "code": code,
            "site_type": site_type,
            "biz_partner": biz_partner,
            "contact": contact,
            "phone": phone,
            "address": address,
        })

    # ---- Sheet2 ----
    sheet2 = []
    for row in ws2.iter_rows(min_row=2, values_only=True):
        name = str(row[0]).strip() if row[0] else ""
        code = str(row[1]).strip() if row[1] else ""
        company = str(row[2]).strip() if row[2] else ""
        if name or code:
            sheet2.append({"name": name, "code": code, "company": company})

    wb.close()
    return sheet1, sheet2


async def geocode_amap(
    session: aiohttp.ClientSession,
    address: str,
    retries: int = MAX_RETRIES,
) -> dict | None:
    """高德地理编码 API 调用."""
    city_hint = extract_city_hint(address)
    params = {
        "key": AMAP_KEY,
        "address": address,
        "city": city_hint,
        "output": "JSON",
    }
    for attempt in range(retries):
        try:
            async with session.get(
                "https://restapi.amap.com/v3/geocode/geo",
                params=params,
                timeout=aiohttp.ClientTimeout(total=15),
            ) as resp:
                if resp.status != 200:
                    if attempt < retries - 1:
                        await asyncio.sleep(RETRY_DELAY[attempt])
                        continue
                    return None
                data = await resp.json()
                if data.get("status") != "1" or not data.get("geocodes"):
                    # 尝试不加 city 限制重试
                    if city_hint and attempt == 0:
                        params["city"] = ""
                        continue
                    if attempt < retries - 1:
                        await asyncio.sleep(RETRY_DELAY[attempt])
                        continue
                    return None

                geo = data["geocodes"][0]
                location = geo.get("location", "")
                if not location:
                    if attempt < retries - 1:
                        await asyncio.sleep(RETRY_DELAY[attempt])
                        continue
                    return None

                lng_str, lat_str = location.split(",")
                lng, lat = float(lng_str), float(lat_str)

                # 校验坐标范围
                if not (LNG_MIN <= lng <= LNG_MAX and LAT_MIN <= lat <= LAT_MAX):
                    if attempt < retries - 1:
                        await asyncio.sleep(RETRY_DELAY[attempt])
                        continue
                    return None

                province = normalize_province(geo.get("province", ""))
                city = normalize_city(geo.get("city", "") or geo.get("district", ""))

                return {
                    "province": province,
                    "city": city,
                    "lng": round(lng, 6),
                    "lat": round(lat, 6),
                    "formatted_address": geo.get("formatted_address", ""),
                }
        except Exception:
            if attempt < retries - 1:
                await asyncio.sleep(RETRY_DELAY[attempt])
                continue
            return None
    return None


async def batch_geocode(sites: list[dict]) -> dict[str, dict]:
    """批量地理编码，返回 {address: geocode_result} 的字典.
    先尝试加载已有的 geocode_results.json 做断点续传.
    """
    cache_file = OUTPUT_DIR / "geocode_results.json"
    cache: dict[str, dict] = {}
    if cache_file.exists():
        try:
            with open(cache_file, "r", encoding="utf-8") as f:
                loaded = json.load(f)
            if isinstance(loaded, dict):
                cache = loaded
                print(f"加载已有地理编码缓存: {len(cache)} 条")
            else:
                print(f"旧版缓存格式（list），将重新生成")
                # 备份旧文件
                old = OUTPUT_DIR / "geocode_results.json.old"
                cache_file.rename(old)
        except Exception:
            pass

    # 去重地址
    unique_addresses = list(dict.fromkeys(
        s["address"] for s in sites if s["address"]
    ))

    todo = [a for a in unique_addresses if a not in cache]
    print(f"总地址: {len(unique_addresses)}, 已缓存: {len(cache)}, 待编码: {len(todo)}")

    if not todo:
        return cache

    async with aiohttp.ClientSession() as session:
        for i, addr in enumerate(todo):
            print(f"  [{i+1}/{len(todo)}] {addr[:60]}...", end=" ", flush=True)
            result = await geocode_amap(session, addr)
            if result:
                cache[addr] = result
                print(f"-> ({result['lng']}, {result['lat']}) {result['province']} {result['city']}")
            else:
                cache[addr] = None
                print("-> [FAILED]")

            # 每 20 条保存一次缓存
            if (i + 1) % 20 == 0:
                with open(cache_file, "w", encoding="utf-8") as f:
                    json.dump(cache, f, ensure_ascii=False, indent=2)

            await asyncio.sleep(BATCH_DELAY)

    # 最终保存
    with open(cache_file, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)

    return cache


def merge_site_with_geocode(site: dict, geocode_cache: dict[str, dict]) -> dict:
    """将站点信息与地理编码结果合并."""
    addr = site["address"]
    geo = geocode_cache.get(addr) if addr else None

    result = {
        "name": site["name"],
        "code": site["code"],
        "company_type": "self" if site["site_type"] == "自营" else "outsourced",
        "parent_name": site["parent_name"],
        "parent_code": make_code(site["parent_name"]),
        "address": addr,
        "province": geo["province"] if geo else "",
        "city": geo["city"] if geo else "",
        "lng": geo["lng"] if geo else 0.0,
        "lat": geo["lat"] if geo else 0.0,
        "geocode_ok": geo is not None,
    }
    return result


async def upsert_repair_sites(conn: asyncpg.Connection, sites: list[dict]) -> dict:
    """更新 repair_network_sites 表. 返回统计."""
    stats = {"inserted": 0, "updated": 0, "failed": 0, "skipped": 0}

    for site in sites:
        if not site["code"]:
            stats["skipped"] += 1
            continue

        try:
            existing = await conn.fetchrow(
                "SELECT id FROM repair_network_sites WHERE code = $1 AND name = $2",
                site["code"], site["name"],
            )
            if existing:
                await conn.execute(
                    """
                    UPDATE repair_network_sites SET
                        name = $1, company_type = $2, parent_name = $3, parent_code = $4,
                        province = $5, city = $6, lng = $7, lat = $8, status = 'active'
                    WHERE id = $9
                    """,
                    site["name"], site["company_type"], site["parent_name"], site["parent_code"],
                    site["province"], site["city"], site["lng"], site["lat"],
                    existing["id"],
                )
                stats["updated"] += 1
            else:
                await conn.execute(
                    """
                    INSERT INTO repair_network_sites
                        (name, code, company_type, parent_name, parent_code,
                         province, city, lng, lat, status)
                    VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, 'active')
                    """,
                    site["name"], site["code"], site["company_type"], site["parent_name"],
                    site["parent_code"], site["province"], site["city"],
                    site["lng"], site["lat"],
                )
                stats["inserted"] += 1
        except Exception as e:
            print(f"  DB Error [{site['code']} {site['name']}]: {e}")
            stats["failed"] += 1

    return stats


async def upsert_yards(conn: asyncpg.Connection, sites: list[dict]) -> dict:
    """更新 yards 表（仅自营网点）. 返回统计."""
    stats = {"inserted": 0, "updated": 0, "failed": 0}

    self_sites = [s for s in sites if s["company_type"] == "self"]

    for site in self_sites:
        if not site["code"]:
            continue
        try:
            existing = await conn.fetchrow(
                "SELECT id FROM yards WHERE code = $1", site["code"],
            )
            if existing:
                await conn.execute(
                    """
                    UPDATE yards SET
                        name = $1, province = $2, city = $3, lng = $4, lat = $5,
                        yard_type = 'yard', status = 'normal'
                    WHERE id = $6
                    """,
                    site["name"], site["province"], site["city"],
                    site["lng"], site["lat"], existing["id"],
                )
                stats["updated"] += 1
            else:
                await conn.execute(
                    """
                    INSERT INTO yards
                        (name, code, yard_type, province, city, lng, lat, capacity, status)
                    VALUES ($1, $2, 'yard', $3, $4, $5, $6, 0, 'normal')
                    """,
                    site["name"], site["code"], site["province"], site["city"],
                    site["lng"], site["lat"],
                )
                stats["inserted"] += 1
        except Exception as e:
            print(f"  Yards DB Error [{site['code']} {site['name']}]: {e}")
            stats["failed"] += 1

    return stats


def match_sheet2_to_sheet1(sheet2: list[dict], sheet1: list[dict], all_sites: list[dict]) -> list[dict]:
    """将 Sheet2 数据按名称匹配到 Sheet1 和已有数据库记录."""
    results = []
    s1_names = {s["name"]: s for s in sheet1}
    db_names = {s["name"]: s for s in all_sites}

    for s2 in sheet2:
        match = None
        source = ""

        # 先按代码精确匹配 Sheet1
        for s1 in sheet1:
            if s1["code"] == s2["code"]:
                match = s1
                source = "s1_code"
                break

        # 再按名称精确匹配 Sheet1
        if not match and s2["name"] in s1_names:
            match = s1_names[s2["name"]]
            source = "s1_name"

        # 名称模糊匹配 Sheet1
        if not match:
            for s1_name, s1_data in s1_names.items():
                if s2["name"] in s1_name or s1_name in s2["name"]:
                    match = s1_data
                    source = "s1_fuzzy"
                    break

        # 匹配已有数据库
        if not match and s2["name"] in db_names:
            match = db_names[s2["name"]]
            source = "db_name"

        # 按 Sheet2 代码匹配数据库
        if not match:
            for db_name, db_data in db_names.items():
                if db_data.get("code") == s2["code"]:
                    match = db_data
                    source = "db_code"
                    break

        results.append({
            "sheet2": s2,
            "match": match,
            "match_source": source,
        })

    return results


async def main():
    print("=" * 60)
    print("Excel 网点地理数据导入")
    print("=" * 60)

    if not EXCEL_PATH.exists():
        print(f"[ERROR] Excel 文件不存在: {EXCEL_PATH}")
        sys.exit(1)

    # 1. 解析 Excel
    print("\n[1/5] 解析 Excel...")
    sheet1, sheet2 = parse_excel(str(EXCEL_PATH))
    print(f"  Sheet1: {len(sheet1)} 条")
    print(f"  Sheet2: {len(sheet2)} 条")

    # 2. 地理编码
    print("\n[2/5] 高德地理编码...")
    geocode_cache = await batch_geocode(sheet1)
    ok_count = sum(1 for v in geocode_cache.values() if v is not None)
    fail_count = sum(1 for v in geocode_cache.values() if v is None)
    print(f"  成功: {ok_count}, 失败: {fail_count}")

    # 合并数据
    sites = [merge_site_with_geocode(s, geocode_cache) for s in sheet1]

    # 3. 保存失败列表
    failed_sites = [s for s in sites if not s["geocode_ok"]]
    if failed_sites:
        failed_file = OUTPUT_DIR / "geocode_failed.json"
        with open(failed_file, "w", encoding="utf-8") as f:
            json.dump([{
                "code": s["code"], "name": s["name"], "address": s["address"],
                "parent_name": s["parent_name"],
            } for s in failed_sites], f, ensure_ascii=False, indent=2)
        print(f"\n  失败地址已保存到: {failed_file} ({len(failed_sites)} 条)")

    # 4. 更新数据库
    print("\n[3/5] 更新数据库...")
    conn = await asyncpg.connect(DB_URL)
    try:
        # repair_network_sites
        print("\n  --- repair_network_sites ---")
        repair_stats = await upsert_repair_sites(conn, sites)
        print(f"  新增: {repair_stats['inserted']}, 更新: {repair_stats['updated']}, "
              f"失败: {repair_stats['failed']}, 跳过: {repair_stats['skipped']}")

        # yards
        print("\n  --- yards ---")
        yard_stats = await upsert_yards(conn, sites)
        print(f"  新增: {yard_stats['inserted']}, 更新: {yard_stats['updated']}, "
              f"失败: {yard_stats['failed']}")

        # Sheet2 匹配
        print("\n[4/5] 处理 Sheet2 匹配...")
        s2_matches = match_sheet2_to_sheet1(sheet2, sites, sites)
        matched = [m for m in s2_matches if m["match"]]
        unmatched = [m for m in s2_matches if not m["match"]]
        print(f"  匹配成功: {len(matched)}, 无匹配: {len(unmatched)}")

        # 对匹配成功的 Sheet2 记录，将其 code 作为别名更新到对应记录的 parent_code 备注
        for m in matched:
            s2 = m["sheet2"]
            match_site = m["match"]
            if s2["code"] and s2["code"] != match_site.get("code", ""):
                # 尝试用 Sheet2 的 code 再更新一次（换 code）
                existing = await conn.fetchrow(
                    "SELECT id FROM repair_network_sites WHERE code = $1 AND name = $2",
                    s2["code"], match_site["name"],
                )
                if existing:
                    await conn.execute(
                        """
                        UPDATE repair_network_sites SET
                            province = $1, city = $2, lng = $3, lat = $3
                        WHERE id = $4
                        """,
                        match_site["province"], match_site["city"],
                        match_site["lng"], match_site["lat"],
                        existing["id"],
                    )
                print(f"    Sheet2 [{s2['code']}] {s2['name']} <- 匹配 [{match_site.get('code','')}] {match_site['name']} (via {m['match_source']})")

        if unmatched:
            manual_file = OUTPUT_DIR / "manual_review.json"
            manual_data = [{
                "sheet2_name": m["sheet2"]["name"],
                "sheet2_code": m["sheet2"]["code"],
                "company": m["sheet2"]["company"],
            } for m in unmatched]
            # 同时加上地理编码失败的
            manual_data.extend([{
                "sheet1_name": s["name"],
                "sheet1_code": s["code"],
                "address": s["address"],
                "reason": "地理编码失败",
            } for s in failed_sites])
            with open(manual_file, "w", encoding="utf-8") as f:
                json.dump(manual_data, f, ensure_ascii=False, indent=2)
            print(f"\n  人工审核清单已保存到: {manual_file}")
            for m in unmatched:
                print(f"    *** 需人工审核: Sheet2 [{m['sheet2']['code']}] {m['sheet2']['name']}")

    finally:
        await conn.close()

    # 5. 汇总报告
    print("\n[5/5] 汇总报告")
    summary = {
        "sheet1_total": len(sheet1),
        "geocode_ok": ok_count,
        "geocode_failed": fail_count,
        "geocode_failed_sites": [s["name"] for s in failed_sites],
        "repair_sites": repair_stats,
        "yards": yard_stats,
        "sheet2_total": len(sheet2),
        "sheet2_matched": len(matched),
        "sheet2_unmatched": len(unmatched),
        "sheet2_unmatched_items": [
            {"name": m["sheet2"]["name"], "code": m["sheet2"]["code"]}
            for m in unmatched
        ],
    }
    summary_file = OUTPUT_DIR / "import_summary.json"
    with open(summary_file, "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)
    print(f"  汇总报告: {summary_file}")
    print(f"\n完成!")
    print(f"  repair_network_sites: +{repair_stats['inserted']} 新增 / ~{repair_stats['updated']} 更新")
    print(f"  yards: +{yard_stats['inserted']} 新增 / ~{yard_stats['updated']} 更新")
    print(f"  地理编码: {ok_count} 成功 / {fail_count} 失败")
    print(f"  Sheet2: {len(matched)} 匹配 / {len(unmatched)} 待审核")


if __name__ == "__main__":
    asyncio.run(main())
